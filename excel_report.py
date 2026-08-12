from __future__ import annotations

from collections import defaultdict
from datetime import date
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import bot_db
from shifts import fmt_hours_clock, fmt_time, now_local, status_text


NAVY = "0F2744"
BLUE = "1F4E79"
ACCENT = "2E75B6"
GREEN_BG = "C6EFCE"
GREEN_FG = "006100"
RED_BG = "FFC7CE"
RED_FG = "9C0006"
ORANGE_BG = "FCE4D6"
ORANGE_FG = "C65911"
YELLOW_BG = "FFF2CC"
LIGHT = "F2F5F9"
WHITE = "FFFFFF"
GRAY = "667085"
LINE = "D0D7E2"
THIN = Side(style="thin", color=LINE)
MED = Side(style="medium", color=BLUE)


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _font(size: int = 11, bold: bool = False, color: str = NAVY, name: str = "Calibri") -> Font:
    return Font(name=name, size=size, bold=bold, color=color)


def _align(h: str = "left", v: str = "center", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _border() -> Border:
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_sheet(ws, title: str, max_col: int, subtitle: str = "") -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5" if subtitle else "A4"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    title_cell = ws.cell(1, 1, title)
    title_cell.font = _font(18, True, WHITE)
    title_cell.fill = _fill(NAVY)
    title_cell.alignment = _align("center")
    ws.row_dimensions[1].height = 34
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
        sub = ws.cell(2, 1, subtitle)
        sub.font = _font(11, False, WHITE)
        sub.fill = _fill(BLUE)
        sub.alignment = _align("center")
        ws.row_dimensions[2].height = 22


def _header(ws, row: int, values: list[str]) -> None:
    for col, value in enumerate(values, 1):
        cell = ws.cell(row, col, value)
        cell.font = _font(11, True, WHITE)
        cell.fill = _fill(ACCENT)
        cell.alignment = _align("center", wrap=True)
        cell.border = _border()
    last = get_column_letter(len(values))
    ws.auto_filter.ref = f"A{row}:{last}{row}"
    ws.row_dimensions[row].height = 28


def _finish(ws, widths: dict[int, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in ws.iter_rows():
        for cell in row:
            if cell.alignment.vertical != "center":
                cell.alignment = cell.alignment.copy(vertical="center")



def _kpi_box(ws, row: int, col: int, title: str, value: Any, bg: str, fg: str = NAVY) -> None:
    """Oddiy 1 ustunli KPI (Kunlik/Kelganlar uchun)."""
    label = ws.cell(row, col, title)
    label.font = _font(10, True, GRAY)
    label.fill = _fill(bg)
    label.alignment = _align("center")
    label.border = _border()
    val = ws.cell(row + 1, col, value)
    val.font = _font(18, True, fg)
    val.fill = _fill(bg)
    val.alignment = _align("center")
    val.border = _border()
    ws.row_dimensions[row].height = 20
    ws.row_dimensions[row + 1].height = 34


def _kpi_card(ws, row: int, col: int, title: str, value: Any, bg: str, fg: str = NAVY) -> None:
    """2 ustunli keng KPI — matn siqilmaydi."""
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
    ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
    for c in (col, col + 1):
        for r in (row, row + 1):
            cell = ws.cell(r, c)
            cell.fill = _fill(bg)
            cell.border = _border()
    label = ws.cell(row, col, title)
    label.font = _font(10, True, GRAY)
    label.alignment = _align("center")
    val = ws.cell(row + 1, col, value)
    val.font = _font(20, True, fg)
    val.alignment = _align("center")
    ws.row_dimensions[row].height = 20
    ws.row_dimensions[row + 1].height = 36


def _section_title(ws, row: int, text: str, max_col: int, bg: str = BLUE) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row, 1, text)
    cell.font = _font(12, True, WHITE)
    cell.fill = _fill(bg)
    cell.alignment = _align("left")
    ws.row_dimensions[row].height = 26


def _dates(start: date, end: date) -> list[date]:
    return bot_db.list_work_days(start, end)


def _attendance_map(records: list[dict[str, Any]]) -> dict[tuple[int, date], dict[str, Any]]:
    return {(int(row["employee_id"]), row["attendance_date"]): row for row in records}


def _row_for(employee: dict[str, Any], day: date, record: dict[str, Any] | None) -> dict[str, Any]:
    record = record or {}
    arrival_status = record.get("arrival_status") or "ABSENT"
    departure_status = record.get("departure_status") or ""
    return {
        "Sana": day,
        "Hafta": f"{((day - date(day.year, day.month, 1)).days // 7) + 1}-hafta",
        "Filial kodi": employee["branch_code"],
        "Filial nomi": employee["branch_name"],
        "F.I.Sh": employee["full_name"],
        "Lavozim": employee["position"],
        "Smena": f"{employee['shift']}-smena",
        "Keldi": fmt_time(record.get("arrival_at")),
        "Kelish holati": status_text(arrival_status),
        "Ketdi": fmt_time(record.get("departure_at")),
        "Ketish holati": status_text(departure_status) if departure_status else "—",
        "Ishlagan vaqt": (
            fmt_hours_clock(record.get("worked_minutes"))
            if record.get("worked_minutes") is not None
            else "—"
        ),
        "Kechikish (soat)": fmt_hours_clock(record.get("late_minutes") or 0),
        "Erta ketish (soat)": fmt_hours_clock(record.get("early_leave_minutes") or 0),
    }


def _write_rows(ws, rows: list[dict[str, Any]], headers: list[str], start_row: int = 3) -> None:
    for row_idx, row in enumerate(rows, start_row):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row_idx, col_idx, row.get(header, "—"))
            cell.border = _border()
            cell.alignment = _align("center" if header != "F.I.Sh" else "left")
            if header == "Sana" and isinstance(cell.value, date):
                cell.number_format = "dd.mm.yyyy"
            _apply_status_colors(cell, header, row)


def _status_colors(text: str) -> tuple[str | None, str | None]:
    value = str(text or "")
    if value.startswith("🟢") or value.startswith("✅"):
        return GREEN_BG, GREEN_FG
    if value.startswith("🔴") or value.startswith("❌"):
        return RED_BG, RED_FG
    if value.startswith("🟠") or value.startswith("⚠️"):
        return ORANGE_BG, ORANGE_FG
    return None, None


def _is_nonzero_clock(value: Any) -> bool:
    text = str(value or "").strip()
    return bool(text) and text not in {"—", "00:00", "0:00"}


ATTENDANCE_DETAIL_HEADERS = [
    "№",
    "Filial",
    "F.I.Sh",
    "Lavozim",
    "Smena",
    "Keldi",
    "Kelish holati",
    "Ketdi",
    "Ketish holati",
    "Ishlagan vaqt",
    "Kechikish (soat)",
    "Erta ketish (soat)",
]


def _apply_status_colors(cell, header: str, row: dict[str, Any]) -> None:
    """Ustun bo‘yicha tushunarli ranglar: vaqtida / kechikkan / erta ketgan / kelmagan."""
    arrival = str(row.get("Kelish holati", ""))
    departure = str(row.get("Ketish holati", ""))

    if header in {"Kelish holati", "Holat"}:
        bg, fg = _status_colors(arrival)
        if bg:
            cell.fill = _fill(bg)
            cell.font = _font(11, True, fg or NAVY)
        return

    if header == "Ketish holati":
        bg, fg = _status_colors(departure)
        if bg:
            cell.fill = _fill(bg)
            cell.font = _font(11, True, fg or NAVY)
        return

    if header == "Kechikish (soat)" and (_is_nonzero_clock(row.get(header)) or arrival.startswith("🔴")):
        cell.fill = _fill(RED_BG)
        cell.font = _font(11, True, RED_FG)
        return

    if header == "Erta ketish (soat)" and (
        _is_nonzero_clock(row.get(header)) or departure.startswith("🟠")
    ):
        cell.fill = _fill(ORANGE_BG)
        cell.font = _font(11, True, ORANGE_FG)
        return

    if header in {"Keldi", "Ketdi", "Ishlagan vaqt"}:
        if arrival.startswith("❌"):
            cell.fill = _fill(RED_BG)
        elif arrival.startswith("🔴"):
            cell.fill = _fill(RED_BG)
        elif arrival.startswith("🟢"):
            cell.fill = _fill(GREEN_BG)


def _write_attendance_detail(
    ws,
    start_row: int,
    rows: list[dict[str, Any]],
    *,
    include_sana: bool = False,
    focus_day: date | None = None,
) -> int:
    """To‘liq davomat ustunlari bilan jadval yozadi. Keyingi bo‘sh qator indeksini qaytaradi."""
    headers = (["№", "Sana"] + ATTENDANCE_DETAIL_HEADERS[1:]) if include_sana else list(ATTENDANCE_DETAIL_HEADERS)
    _header(ws, start_row, headers)
    if not rows:
        return start_row + 2

    for idx, row in enumerate(rows, 1):
        r = start_row + idx
        values: list[Any] = [idx]
        if include_sana:
            values.append(focus_day or row.get("Sana"))
        values.extend(
            [
                row["Filial nomi"],
                row["F.I.Sh"],
                row["Lavozim"],
                row["Smena"],
                row["Keldi"],
                row["Kelish holati"],
                row["Ketdi"],
                row["Ketish holati"],
                row["Ishlagan vaqt"],
                row["Kechikish (soat)"],
                row["Erta ketish (soat)"],
            ]
        )
        late = str(row["Kelish holati"]).startswith("🔴")
        absent = str(row["Kelish holati"]).startswith("❌")
        for col, (header, value) in enumerate(zip(headers, values), 1):
            cell = ws.cell(r, col, value)
            cell.border = _border()
            cell.alignment = _align("left" if header == "F.I.Sh" else "center")
            cell.font = _font(11, False, NAVY)
            if header == "Sana" and isinstance(value, date):
                cell.number_format = "dd.mm.yyyy"
            # Qator foni: kechikkan qizil, vaqtida yashil, kelmagan qizil
            if absent:
                cell.fill = _fill(RED_BG)
            elif late:
                cell.fill = _fill(RED_BG)
            else:
                cell.fill = _fill(GREEN_BG)
            _apply_status_colors(cell, header, row)
    return start_row + len(rows) + 2


def _build_dashboard(
    wb: Workbook,
    label: str,
    focus_day: date,
    _daily_rows: list[dict[str, Any]],
    employees: list[dict[str, Any]],
    day_rows: list[dict[str, Any]],
) -> None:
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A7"

    for col in range(1, 13):
        ws.cell(1, col).fill = _fill(NAVY)
        ws.cell(2, col).fill = _fill(BLUE)
    ws.merge_cells("A1:L1")
    title = ws.cell(1, 1, "FILIAL ATTENDANCE")
    title.font = _font(22, True, WHITE)
    title.fill = _fill(NAVY)
    title.alignment = _align("center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:L2")
    sub = ws.cell(
        2,
        1,
        f"Sana: {focus_day.strftime('%d.%m.%Y')}     ·     Davr: {label}     ·     {now_local().strftime('%d.%m.%Y %H:%M')}",
    )
    sub.font = _font(11, False, WHITE)
    sub.fill = _fill(BLUE)
    sub.alignment = _align("center")
    ws.row_dimensions[2].height = 24

    arrived = sum(1 for row in day_rows if row["Keldi"] != "—")
    late = sum(1 for row in day_rows if row["Kelish holati"].startswith("🔴"))
    absent = sum(1 for row in day_rows if row["Kelish holati"].startswith("❌"))
    on_time = sum(1 for row in day_rows if row["Kelish holati"].startswith("🟢"))
    total_emp = len(employees)
    rate = (arrived / total_emp * 100) if total_emp else 0.0

    ws.merge_cells("A4:L4")
    head = ws.cell(4, 1, f"BUGUNGI KO‘RSATKICHLAR — {focus_day.strftime('%d.%m.%Y')}")
    head.font = _font(13, True, WHITE)
    head.fill = _fill(NAVY)
    head.alignment = _align("left")
    ws.row_dimensions[4].height = 28

    cards = [
        (1, "JAMI", total_emp, LIGHT, NAVY),
        (3, "KELGAN", arrived, GREEN_BG, GREEN_FG),
        (5, "VAQTIDA", on_time, GREEN_BG, GREEN_FG),
        (7, "KECHIKKAN", late, ORANGE_BG, ORANGE_FG),
        (9, "KELMAGAN", absent, RED_BG, RED_FG),
        (11, "DAVOMAT", f"{rate:.1f}%", YELLOW_BG, NAVY),
    ]
    for col, title_text, value, bg, fg in cards:
        _kpi_card(ws, 5, col, title_text, value, bg, fg)

    _finish(
        ws,
        {
            1: 11,
            2: 11,
            3: 11,
            4: 11,
            5: 11,
            6: 11,
            7: 12,
            8: 12,
            9: 11,
            10: 11,
            11: 12,
            12: 12,
        },
    )


def _detail_widths(*, with_sana: bool = False) -> dict[int, int]:
    base = {
        1: 6,
        2: 24,
        3: 26,
        4: 16,
        5: 11,
        6: 10,
        7: 16,
        8: 10,
        9: 16,
        10: 14,
        11: 14,
        12: 14,
    }
    if not with_sana:
        return base
    return {
        1: 6,
        2: 12,
        3: 24,
        4: 26,
        5: 16,
        6: 11,
        7: 10,
        8: 16,
        9: 10,
        10: 16,
        11: 14,
        12: 14,
        13: 14,
    }


def _build_kunlik(
    wb: Workbook,
    focus_day: date,
    day_rows: list[dict[str, Any]],
) -> None:
    ws = wb.create_sheet("Kunlik", 1)
    date_label = focus_day.strftime("%d.%m.%Y")
    cols = len(ATTENDANCE_DETAIL_HEADERS)
    _style_sheet(
        ws,
        f"KUNLIK HISOBOT — {date_label}",
        cols,
        subtitle=f"Sana: {date_label}   |   Ish kuni: {'Ha' if bot_db.is_work_day(focus_day) else 'Yo‘q'}",
    )

    arrived_rows = [row for row in day_rows if row["Keldi"] != "—"]
    absent_rows = [row for row in day_rows if row["Kelish holati"].startswith("❌")]
    late_rows = [row for row in day_rows if row["Kelish holati"].startswith("🔴")]

    _kpi_box(ws, 4, 1, "SANA", date_label, LIGHT, NAVY)
    _kpi_box(ws, 4, 2, "JAMI", len(day_rows), LIGHT, NAVY)
    _kpi_box(ws, 4, 3, "KELGAN", len(arrived_rows), GREEN_BG, GREEN_FG)
    _kpi_box(ws, 4, 4, "KECHIKKAN", len(late_rows), ORANGE_BG, ORANGE_FG)
    _kpi_box(ws, 4, 5, "KELMAGAN", len(absent_rows), RED_BG, RED_FG)

    start = 7
    _section_title(ws, start, f"✅ BUGUN KELGANLAR — {date_label}", cols, GREEN_FG)
    if arrived_rows:
        next_row = _write_attendance_detail(ws, start + 1, arrived_rows)
    else:
        ws.cell(start + 2, 1, "Bugun kelganlar yo‘q.").font = _font(11, False, GRAY)
        next_row = start + 4

    _section_title(ws, next_row, f"❌ BUGUN KELMAGANLAR — {date_label}", cols, RED_FG)
    if absent_rows:
        _write_attendance_detail(ws, next_row + 1, absent_rows)
    else:
        msg = (
            "Bugun barcha xodimlar kelgan."
            if bot_db.is_work_day(focus_day)
            else "Bugun ish kuni emas — kelmaganlar hisoblanmaydi."
        )
        ws.cell(next_row + 2, 1, msg).font = _font(11, False, GRAY)

    _finish(ws, _detail_widths())


def _build_kelganlar(wb: Workbook, focus_day: date, day_rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Kelganlar", 2)
    date_label = focus_day.strftime("%d.%m.%Y")
    arrived_rows = [row for row in day_rows if row["Keldi"] != "—"]
    cols = len(ATTENDANCE_DETAIL_HEADERS) + 1  # + Sana
    _style_sheet(
        ws,
        f"BUGUN KELGANLAR — {date_label}",
        cols,
        subtitle=f"Sana: {date_label}   |   Jami kelgan: {len(arrived_rows)} ta",
    )

    _kpi_box(ws, 4, 1, "SANA", date_label, LIGHT, NAVY)
    _kpi_box(ws, 4, 2, "KELGANLAR", len(arrived_rows), GREEN_BG, GREEN_FG)
    on_time = sum(1 for row in arrived_rows if row["Kelish holati"].startswith("🟢"))
    late = sum(1 for row in arrived_rows if row["Kelish holati"].startswith("🔴"))
    _kpi_box(ws, 4, 3, "VAQTIDA", on_time, GREEN_BG, GREEN_FG)
    _kpi_box(ws, 4, 4, "KECHIKKAN", late, ORANGE_BG, ORANGE_FG)

    if arrived_rows:
        _write_attendance_detail(ws, 7, arrived_rows, include_sana=True, focus_day=focus_day)
    else:
        _header(ws, 7, ["№", "Sana"] + ATTENDANCE_DETAIL_HEADERS[1:])
        ws.cell(8, 1, f"{date_label} sanasida kelganlar yo‘q.").font = _font(11, False, GRAY)

    _finish(ws, _detail_widths(with_sana=True))


def build_report(start: date, end: date, label: str) -> BytesIO:
    employees = bot_db.get_all_employees()
    records = bot_db.get_attendance_range(start, end)
    record_map = _attendance_map(records)
    days = _dates(start, end)
    daily_rows = [
        _row_for(employee, day, record_map.get((int(employee["id"]), day)))
        for day in days
        for employee in employees
    ]

    focus_day = end
    # Kunlik fokusat focus_day; agar ish kuni bo‘lmasa ham xodimlar holatini ko‘rsatamiz
    day_rows = [
        _row_for(employee, focus_day, record_map.get((int(employee["id"]), focus_day)))
        for employee in employees
    ]

    wb = Workbook()
    _build_dashboard(wb, label, focus_day, daily_rows, employees, day_rows)
    _build_kunlik(wb, focus_day, day_rows)
    _build_kelganlar(wb, focus_day, day_rows)

    monthly = wb.create_sheet("Oylik hisobot")
    monthly_headers = [
        "№", "Filial kodi", "Filial nomi", "F.I.Sh", "Lavozim", "Smena",
        "Ish kunlari", "Kelgan", "Kelmagan", "Kechikkan", "Jami kechikish (soat)",
        "Jami ishlagan (soat)", "Erta ketish (soat)",
    ]
    _style_sheet(monthly, f"Oylik hisobot: {label}", len(monthly_headers), subtitle=f"Davr: {label}")
    _header(monthly, 3, monthly_headers)
    stats = bot_db.get_monthly_stats(start, end)
    for idx, row in enumerate(stats, 1):
        days_count = len(days)
        values = [
            idx, row["branch_code"], row["branch_name"], row["full_name"], row["position"],
            f"{row['shift']}-smena", days_count, row["arrived"],
            max(0, days_count - row["arrived"]), row["late_days"],
            fmt_hours_clock(row["late_minutes"]),
            fmt_hours_clock(row["worked_minutes"]),
            fmt_hours_clock(row["early_minutes"]),
        ]
        for col, value in enumerate(values, 1):
            cell = monthly.cell(idx + 3, col, value)
            cell.border = _border()
    _finish(monthly, {1: 6, 2: 16, 3: 22, 4: 24, 5: 18, 6: 12, 7: 12, 8: 10, 9: 12, 10: 12, 11: 20, 12: 20, 13: 18})

    daily_headers = [
        "Sana", "Hafta", "Filial kodi", "Filial nomi", "F.I.Sh", "Lavozim", "Smena",
        "Keldi", "Kelish holati", "Ketdi", "Ketish holati", "Ishlagan vaqt",
        "Kechikish (soat)", "Erta ketish (soat)",
    ]
    for week_index, week_start in enumerate(range(0, len(days), 7), 1):
        week_days = days[week_start : week_start + 7]
        ws = wb.create_sheet(f"{week_index}-hafta")
        _style_sheet(ws, f"{week_index}-HAFTA", len(daily_headers))
        _header(ws, 3, daily_headers)
        rows = [row for row in daily_rows if row["Sana"] in week_days]
        _write_rows(ws, rows, daily_headers, start_row=4)
        _finish(ws, {1: 13, 2: 11, 3: 15, 4: 22, 5: 24, 6: 17, 7: 11, 8: 10, 9: 20, 10: 10, 11: 20, 12: 18, 13: 16, 14: 18})

    late_ws = wb.create_sheet("Kechikishlar")
    late_headers = ["Sana", "Filial", "F.I.Sh", "Lavozim", "Smena", "Belgilangan", "Keldi", "Kechikish"]
    _style_sheet(late_ws, "Kechikishlar", len(late_headers))
    _header(late_ws, 3, late_headers)
    late_rows = []
    for row in daily_rows:
        if row["Kelish holati"].startswith("🔴"):
            late_rows.append({
                "Sana": row["Sana"], "Filial": row["Filial nomi"], "F.I.Sh": row["F.I.Sh"],
                "Lavozim": row["Lavozim"], "Smena": row["Smena"],
                "Belgilangan": "08:15" if row["Smena"].startswith("1") else "17:15",
                "Keldi": row["Keldi"], "Kechikish": row["Kechikish (soat)"],
            })
    _write_rows(late_ws, late_rows, late_headers, start_row=4)
    _finish(late_ws, {1: 13, 2: 22, 3: 24, 4: 18, 5: 11, 6: 15, 7: 12, 8: 16})

    absent_ws = wb.create_sheet("Kelmaganlar")
    absent_headers = ["Sana", "Filial", "F.I.Sh", "Lavozim", "Smena", "Holat"]
    _style_sheet(absent_ws, "Kelmaganlar", len(absent_headers))
    _header(absent_ws, 3, absent_headers)
    absent_rows = [
        {
            "Sana": row["Sana"], "Filial": row["Filial nomi"], "F.I.Sh": row["F.I.Sh"],
            "Lavozim": row["Lavozim"], "Smena": row["Smena"], "Holat": "❌ Kelmagan",
        }
        for row in daily_rows
        if row["Kelish holati"].startswith("❌")
    ]
    _write_rows(absent_ws, absent_rows, absent_headers, start_row=4)
    _finish(absent_ws, {1: 13, 2: 22, 3: 24, 4: 18, 5: 11, 6: 18})

    branches_ws = wb.create_sheet("Filiallar")
    branch_headers = ["№", "Filial kodi", "Filial nomi", "Xodimlar", "Kelganlar", "Kechikkanlar", "Kelmaganlar", "Davomat"]
    _style_sheet(branches_ws, "Filiallar statistikasi", len(branch_headers))
    _header(branches_ws, 3, branch_headers)
    branch_groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in daily_rows:
        branch_groups[row["Filial kodi"]].append(row)
    for idx, (code, rows) in enumerate(branch_groups.items(), 1):
        present_count = sum(1 for row in rows if row["Keldi"] != "—")
        late_count = sum(1 for row in rows if row["Kelish holati"].startswith("🔴"))
        absent_count = sum(1 for row in rows if row["Kelish holati"].startswith("❌"))
        values = [
            idx, code, rows[0]["Filial nomi"], len({r["F.I.Sh"] for r in rows}),
            present_count, late_count, absent_count,
            f"{present_count / len(rows) * 100:.1f}%" if rows else "0.0%",
        ]
        for col, value in enumerate(values, 1):
            cell = branches_ws.cell(idx + 3, col, value)
            cell.border = _border()
    _finish(branches_ws, {1: 6, 2: 16, 3: 24, 4: 12, 5: 12, 6: 15, 7: 14, 8: 12})

    employees_ws = wb.create_sheet("Xodimlar")
    employee_headers = ["№", "Filial kodi", "Filial nomi", "F.I.Sh", "Telefon", "Lavozim", "Smena", "Holat"]
    _style_sheet(employees_ws, "Xodimlar", len(employee_headers))
    _header(employees_ws, 3, employee_headers)
    for idx, employee in enumerate(employees, 1):
        values = [
            idx, employee["branch_code"], employee["branch_name"], employee["full_name"],
            employee["phone"], employee["position"], f"{employee['shift']}-smena", "Faol",
        ]
        for col, value in enumerate(values, 1):
            cell = employees_ws.cell(idx + 3, col, value)
            cell.border = _border()
    _finish(employees_ws, {1: 6, 2: 16, 3: 24, 4: 24, 5: 18, 6: 18, 7: 11, 8: 12})

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "❌ Kelmagan":
                    cell.fill = _fill(RED_BG)
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def build_list_report(kind: str) -> BytesIO:
    """Bitta listli Excel: employees | late | absent."""
    today = now_local().date()
    date_label = today.strftime("%d.%m.%Y")

    if kind == "employees":
        rows = bot_db.get_all_employees()
        title = "XODIMLAR RO‘YXATI"
        sheet_name = "Xodimlar"
        headers = ["№", "Filial kodi", "Filial", "F.I.Sh", "Telefon", "Lavozim", "Smena"]
        data = [
            [
                idx,
                row.get("branch_code", ""),
                row.get("branch_name", ""),
                row.get("full_name", ""),
                row.get("phone") or "—",
                row.get("position") or "—",
                f"{row.get('shift')}-smena",
            ]
            for idx, row in enumerate(rows, 1)
        ]
        widths = {1: 6, 2: 14, 3: 28, 4: 26, 5: 16, 6: 18, 7: 12}
        empty_msg = "Hozircha xodimlar ro‘yxatdan o‘tmagan."
    elif kind == "late":
        rows = [row for row in bot_db.get_today_rows(today) if row.get("arrival_status") == "LATE"]
        title = f"KECHIKKANLAR — {date_label}"
        sheet_name = "Kechikkanlar"
        headers = ["№", "Filial", "F.I.Sh", "Lavozim", "Smena", "Keldi", "Kechikish"]
        data = [
            [
                idx,
                row.get("branch_name", ""),
                row.get("full_name", ""),
                row.get("position") or "—",
                f"{row.get('shift')}-smena",
                fmt_time(row.get("arrival_at")),
                fmt_hours_clock(row.get("late_minutes") or 0),
            ]
            for idx, row in enumerate(rows, 1)
        ]
        widths = {1: 6, 2: 28, 3: 26, 4: 18, 5: 12, 6: 12, 7: 12}
        empty_msg = "Bugun kechikkanlar yo‘q."
    elif kind == "absent":
        rows = [row for row in bot_db.get_today_rows(today) if not row.get("arrival_at")]
        title = f"KELMAGANLAR — {date_label}"
        sheet_name = "Kelmaganlar"
        headers = ["№", "Filial", "F.I.Sh", "Lavozim", "Smena", "Holat"]
        data = [
            [
                idx,
                row.get("branch_name", ""),
                row.get("full_name", ""),
                row.get("position") or "—",
                f"{row.get('shift')}-smena",
                "Kelmagan",
            ]
            for idx, row in enumerate(rows, 1)
        ]
        widths = {1: 6, 2: 28, 3: 26, 4: 18, 5: 12, 6: 14}
        empty_msg = "Bugun kelmaganlar yo‘q."
    else:
        raise ValueError(f"Unknown list kind: {kind}")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    _style_sheet(
        ws,
        title,
        len(headers),
        subtitle=f"Sana: {date_label}   ·   Jami: {len(data)} ta",
    )
    _header(ws, 3, headers)

    if not data:
        cell = ws.cell(4, 1, empty_msg)
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=len(headers))
        cell.font = _font(11, False, GRAY)
        cell.alignment = _align("left")
    else:
        name_col = headers.index("F.I.Sh") + 1
        for r_idx, values in enumerate(data, 4):
            for c_idx, value in enumerate(values, 1):
                cell = ws.cell(r_idx, c_idx, value)
                cell.border = _border()
                cell.alignment = _align("left" if c_idx == name_col else "center")
                cell.font = _font(11, False, NAVY)
                if r_idx % 2 == 0:
                    cell.fill = _fill(LIGHT)
                if kind == "late" and c_idx == 7:
                    cell.fill = _fill(ORANGE_BG)
                    cell.font = _font(11, True, ORANGE_FG)
                if kind == "absent" and c_idx == 6:
                    cell.fill = _fill(RED_BG)
                    cell.font = _font(11, True, RED_FG)

    _finish(ws, widths)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
