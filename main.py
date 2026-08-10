from __future__ import annotations

import asyncio
import logging
import os
from datetime import date, datetime, timedelta
from io import BytesIO

from openpyxl import load_workbook
from telegram import (
    BotCommand,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    InputFile,
    KeyboardButton,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
    Update,
)
from telegram.ext import (
    Application,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    filters,
)

import bot_db
from excel_report import build_report
from shifts import (
    LOCAL_TZ,
    arrival_result,
    departure_result,
    fmt_duration,
    fmt_time,
    now_local,
    status_text,
    worked_minutes,
)


logging.basicConfig(
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger("filial-attendance")
# Telegram client URLs may contain the bot token. Keep transport logs quiet so
# secrets never end up in workflow output.
logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("httpcore").setLevel(logging.WARNING)

EMPLOYEE_MENU = [
    ["🟢 KELDIM", "🔴 KETDIM"],
    ["📊 Bugungi holatim", "📅 Davomatim"],
    ["👤 Profilim"],
]
ADMIN_MENU = [
    ["📊 DASHBOARD", "🏢 FILIALLAR"],
    ["👥 XODIMLAR", "🔴 KECHIKKANLAR"],
    ["❌ KELMAGANLAR", "📥 EXCEL HISOBOT"],
    ["📥 Excel yuklash", "⚙️ SOZLAMALAR"],
    ["📝 So‘rovlar"],
]


def menu_keyboard(rows: list[list[str]]) -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(rows, resize_keyboard=True)


def admin_user(user_id: int) -> bool:
    return bot_db.is_admin(user_id)


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user = update.effective_user
    if not user or not update.message:
        return
    context.user_data.clear()
    if admin_user(user.id):
        await update.message.reply_text(
            "📊 FILIAL ATTENDANCE admin paneliga xush kelibsiz.",
            reply_markup=menu_keyboard(ADMIN_MENU),
        )
        return
    employee = bot_db.get_employee(user.id)
    if employee:
        await update.message.reply_text(
            f"Assalomu alaykum, {employee['full_name']}!\n"
            "Kerakli amalni tanlang.",
            reply_markup=menu_keyboard(EMPLOYEE_MENU),
        )
        return
    buttons = [[InlineKeyboardButton("📝 Ro‘yxatdan o‘tish", callback_data="register")]]
    await update.message.reply_text(
        "👋 Assalomu alaykum!\n\n"
        "FILIAL ATTENDANCE tizimiga xush kelibsiz.\n"
        "Davomatdan foydalanish uchun ro‘yxatdan o‘ting.",
        reply_markup=InlineKeyboardMarkup(buttons),
    )


async def whoami(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.message and update.effective_user:
        await update.message.reply_text(
            f"Sizning Telegram ID raqamingiz: {update.effective_user.id}\n"
            "Admin qilish uchun shu ID ni bot sozlamasiga qo‘shing."
        )


async def callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    user = update.effective_user
    if not query or not user:
        return
    await query.answer()
    data = query.data or ""

    if data == "register":
        context.user_data["registration_step"] = "phone"
        phone_button = KeyboardButton("📱 Telefon raqamimni yuborish", request_contact=True)
        await query.message.reply_text(
            "Telefon raqamingizni yuboring:",
            reply_markup=ReplyKeyboardMarkup([[phone_button]], resize_keyboard=True, one_time_keyboard=True),
        )
        return

    if data == "branch_search":
        context.user_data["registration_step"] = "branch_search"
        await query.message.reply_text(
            "🔎 Filial kodi yoki nomini yozing:",
            reply_markup=ReplyKeyboardRemove(),
        )
        return

    if data.startswith("branch:"):
        if context.user_data.get("registration_step") not in {"branch", "branch_search"}:
            return
        branch_id = int(data.split(":", 1)[1])
        branch = bot_db.get_branch(branch_id)
        if not branch:
            await query.message.reply_text("❌ Filial topilmadi.")
            return
        context.user_data["branch_id"] = branch_id
        context.user_data["registration_step"] = "position"
        keyboard = InlineKeyboardMarkup(
            [
                [InlineKeyboardButton("👔 Boshqaruvchi", callback_data="position:Boshqaruvchi")],
                [InlineKeyboardButton("💊 Farmatsevt", callback_data="position:Farmatsevt")],
            ]
        )
        await query.message.reply_text(
            f"🏢 Tanlangan filial: {branch['branch_name']}\n\n"
            "👤 Lavozimingizni tanlang:",
            reply_markup=keyboard,
        )
        return

    if data.startswith("position:"):
        if context.user_data.get("registration_step") != "position":
            return
        context.user_data["position"] = data.split(":", 1)[1]
        context.user_data["registration_step"] = "shift"
        keyboard = InlineKeyboardMarkup(
            [
                [InlineKeyboardButton("1️⃣ 1-SMENA", callback_data="shift:1")],
                [InlineKeyboardButton("2️⃣ 2-SMENA", callback_data="shift:2")],
            ]
        )
        await query.message.reply_text("🕐 Ish smenangizni tanlang:", reply_markup=keyboard)
        return

    if data.startswith("shift:"):
        if context.user_data.get("registration_step") != "shift":
            return
        shift = data.split(":", 1)[1]
        employee = bot_db.create_employee(
            telegram_id=user.id,
            phone=context.user_data["phone"],
            full_name=context.user_data["full_name"],
            branch_id=context.user_data["branch_id"],
            position=context.user_data["position"],
            shift=shift,
        )
        context.user_data.clear()
        await query.message.reply_text(
            "✅ Ro‘yxatdan o‘tish yakunlandi.\n\n"
            f"👤 F.I.Sh: {employee['full_name']}\n"
            f"🏢 Filial: {employee['branch_name']}\n"
            f"👔 Lavozim: {employee['position']}\n"
            f"🕐 Smena: {shift}-SMENA",
            reply_markup=menu_keyboard(EMPLOYEE_MENU),
        )
        return

    if data.startswith("report:"):
        if not admin_user(user.id):
            return
        kind = data.split(":", 1)[1]
        await send_report(query.message, kind)
        return

    if data.startswith("branch_view:"):
        if not admin_user(user.id):
            return
        branch_id = int(data.split(":", 1)[1])
        await show_branch(query.message, branch_id)
        return


async def contact(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not update.message or not update.effective_user:
        return
    if context.user_data.get("registration_step") != "phone":
        return
    received = update.message.contact
    if not received or received.user_id not in (None, update.effective_user.id):
        await update.message.reply_text("❌ Faqat o‘zingizning telefon raqamingizni yuboring.")
        return
    context.user_data["phone"] = received.phone_number
    context.user_data["registration_step"] = "name"
    await update.message.reply_text(
        "F.I.Sh.ni kiriting:\nMasalan: Aliyev Vali",
        reply_markup=ReplyKeyboardRemove(),
    )


async def show_branches(message, search: str = "") -> None:
    branches = bot_db.get_branches(search)
    if not branches:
        await message.reply_text("❌ Mos filial topilmadi. Boshqa nom bilan urinib ko‘ring.")
        return
    buttons = [
        [InlineKeyboardButton(f"🏢 {branch['branch_name']}", callback_data=f"branch:{branch['id']}")]
        for branch in branches[:80]
    ]
    buttons.append([InlineKeyboardButton("🔎 Filialni qidirish", callback_data="branch_search")])
    await message.reply_text(
        "🏢 Filialingizni tanlang:",
        reply_markup=InlineKeyboardMarkup(buttons),
    )


async def document(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not update.message or not update.effective_user or not admin_user(update.effective_user.id):
        return
    if context.user_data.get("admin_state") != "branch_excel":
        await update.message.reply_text("Avval “📥 Excel yuklash” tugmasini bosing.")
        return
    document_file = update.message.document
    if not document_file or not document_file.file_name.lower().endswith((".xlsx", ".xlsm")):
        await update.message.reply_text("❌ Faqat .xlsx yoki .xlsm formatdagi Excel fayl yuboring.")
        return
    file = await document_file.get_file()
    buffer = BytesIO()
    await file.download_to_memory(buffer)
    buffer.seek(0)
    try:
        workbook = load_workbook(buffer, read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            raise ValueError("Fayl bo‘sh.")
        start = 1 if len(values[0]) >= 2 and str(values[0][0]).strip().lower() in {"filial kodi", "kod", "code"} else 0
        rows: list[tuple[str, str]] = []
        errors: list[str] = []
        for row_number, row in enumerate(values[start:], start + 1):
            code = str(row[0] or "").strip() if len(row) > 0 else ""
            name = str(row[1] or "").strip() if len(row) > 1 else ""
            if not code and not name:
                continue
            if not code:
                errors.append(f"{row_number}-qator: Filial kodi mavjud emas.")
            elif not name:
                errors.append(f"{row_number}-qator: Filial nomi mavjud emas.")
            else:
                rows.append((code, name))
        if errors:
            await update.message.reply_text("❌ Excel faylida xatolik bor:\n" + "\n".join(errors[:20]) + "\n\nFayl qabul qilinmadi.")
            return
        added, changed = bot_db.import_branches(rows)
        context.user_data.pop("admin_state", None)
        await update.message.reply_text(
            "✅ Fayl muvaffaqiyatli qabul qilindi.\n\n"
            f"🏢 Filiallar: {len(rows)} ta\n"
            f"Yangi qo‘shilgan: {added}\n"
            f"Yangilangan: {changed}\n"
            "Xatolik: 0",
            reply_markup=menu_keyboard(ADMIN_MENU),
        )
    except Exception as exc:
        logger.exception("Excel import error: %s", exc)
        await update.message.reply_text("❌ Excel faylini o‘qib bo‘lmadi. A va B ustunlarini tekshiring.")


async def text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not update.message or not update.effective_user:
        return
    text_value = (update.message.text or "").strip()
    user_id = update.effective_user.id

    if context.user_data.get("registration_step") == "name":
        if len(text_value) < 3:
            await update.message.reply_text("F.I.Sh.ni to‘liqroq kiriting.")
            return
        context.user_data["full_name"] = text_value
        context.user_data["registration_step"] = "branch"
        await show_branches(update.message)
        return
    if context.user_data.get("registration_step") == "branch_search":
        context.user_data["registration_step"] = "branch"
        await show_branches(update.message, text_value)
        return

    if admin_user(user_id):
        await admin_text(update, context, text_value)
    else:
        await employee_text(update, context, text_value)


async def employee_text(update: Update, context: ContextTypes.DEFAULT_TYPE, text_value: str) -> None:
    employee = bot_db.get_employee(update.effective_user.id)  # type: ignore[union-attr]
    if not employee:
        await update.message.reply_text("Avval /start orqali ro‘yxatdan o‘ting.")  # type: ignore[union-attr]
        return
    if text_value == "🟢 KELDIM":
        current = now_local()
        status, late = arrival_result(current, employee["shift"])
        existing = bot_db.get_attendance_for_date(employee["id"], current.date())
        if existing and existing["arrival_at"] is not None:
            await update.message.reply_text(  # type: ignore[union-attr]
                "⚠️ Bugun sizning kelish vaqtingiz allaqachon qayd etilgan.\n"
                f"🕐 Kelgan vaqt: {fmt_time(existing['arrival_at'])}"
            )
            return
        row = bot_db.record_arrival(employee["id"], current.date(), current, status, late)
        result_line = status_text(status) + (f" ({late} daqiqa)" if late else "")
        await update.message.reply_text(  # type: ignore[union-attr]
            "✅ KELISH QAYD ETILDI\n\n"
            f"🏢 {employee['branch_name']}\n"
            f"🕐 Kelgan vaqt: {fmt_time(current)}\n"
            f"📌 Holat: {result_line}"
        )
        return
    if text_value == "🔴 KETDIM":
        current = now_local()
        row = bot_db.find_open_attendance(employee["id"], current.date())
        if not row:
            await update.message.reply_text("❌ Bugun avval “🟢 KELDIM” tugmasini bosing.")  # type: ignore[union-attr]
            return
        status, early = departure_result(current, employee["shift"])
        duration = worked_minutes(row["arrival_at"], current)
        saved = bot_db.record_departure(row["id"], current, duration, status, early)
        if not saved:
            await update.message.reply_text("⚠️ Ketish vaqti allaqachon qayd etilgan.")  # type: ignore[union-attr]
            return
        early_line = status_text(status) + (f" ({early} daqiqa oldin)" if early else "")
        await update.message.reply_text(  # type: ignore[union-attr]
            "✅ KETISH QAYD ETILDI\n\n"
            f"🕐 Kelgan: {fmt_time(row['arrival_at'])}\n"
            f"🕐 Ketgan: {fmt_time(current)}\n"
            f"⏱ Ishlagan vaqt: {fmt_duration(duration)}\n"
            f"📌 Holat: {early_line}\n\n"
            "🟢 Ish kuni yakunlandi."
        )
        return
    if text_value == "📊 Bugungi holatim":
        await today_status(update.message, employee)
        return
    if text_value == "📅 Davomatim":
        await attendance_summary(update.message, employee)
        return
    if text_value == "👤 Profilim":
        await profile(update.message, employee)
        return
    await update.message.reply_text("Kerakli tugmani tanlang.", reply_markup=menu_keyboard(EMPLOYEE_MENU))  # type: ignore[union-attr]


async def today_status(message, employee: dict) -> None:
    current = now_local()
    row = bot_db.get_attendance_for_date(employee["id"], current.date())
    await message.reply_text(
        f"📊 BUGUNGI HOLATINGIZ\n\n"
        f"📅 {current.strftime('%d.%m.%Y')}\n"
        f"🏢 {employee['branch_name']}\n"
        f"👔 {employee['position']}\n"
        f"🕐 {employee['shift']}-smena\n\n"
        f"🟢 Keldi: {fmt_time(row['arrival_at']) if row else '—'}\n"
        f"🔴 Ketdi: {fmt_time(row['departure_at']) if row else '—'}\n"
        f"⏱ Ishlagan: {fmt_duration(row['worked_minutes']) if row else '—'}\n"
        f"📌 Holat: {status_text(row['arrival_status'] if row and row['arrival_status'] else 'ABSENT')}"
    )


async def attendance_summary(message, employee: dict) -> None:
    current = now_local()
    start = current.date().replace(day=1)
    next_month = (start.replace(day=28) + timedelta(days=4)).replace(day=1)
    end = next_month - timedelta(days=1)
    if end > current.date():
        end = current.date()
    rows = bot_db.get_attendance_range(start, end)
    mine = [row for row in rows if row["employee_id"] == employee["id"]]
    worked = sum(row["worked_minutes"] or 0 for row in mine)
    late = sum(row["late_minutes"] or 0 for row in mine)
    await message.reply_text(
        f"📊 DAVOMATIM\n\n"
        f"📅 {current.strftime('%B %Y')}\n"
        f"📆 Ish kunlari: {(end - start).days + 1}\n"
        f"🟢 Kelgan: {sum(1 for r in mine if r['arrival_at'])}\n"
        f"🔴 Kechikkan: {sum(1 for r in mine if r['arrival_status'] == 'LATE')}\n"
        f"⏱ Jami ishlagan: {fmt_duration(worked)}\n"
        f"⌛ Jami kechikish: {late} daqiqa"
    )


async def profile(message, employee: dict) -> None:
    await message.reply_text(
        f"👤 PROFILIM\n\n"
        f"👤 F.I.Sh: {employee['full_name']}\n"
        f"📱 Telefon: {employee['phone']}\n"
        f"🏢 Filial: {employee['branch_name']}\n"
        f"👔 Lavozim: {employee['position']}\n"
        f"🕐 Smena: {employee['shift']}-smena"
    )


async def admin_text(update: Update, context: ContextTypes.DEFAULT_TYPE, value: str) -> None:
    message = update.message
    if value in {"📥 EXCEL HISOBOT", "📅 Oylik hisobot"}:
        await message.reply_text(
            "Hisobot turini tanlang:",
            reply_markup=InlineKeyboardMarkup(
                [
                    [InlineKeyboardButton("📅 Bugungi", callback_data="report:today")],
                    [InlineKeyboardButton("📆 Haftalik", callback_data="report:weekly")],
                    [InlineKeyboardButton("📅 Oylik", callback_data="report:monthly")],
                ]
            ),
        )
        return
    if value == "📥 Excel yuklash":
        context.user_data["admin_state"] = "branch_excel"
        await message.reply_text(
            "📥 Filiallar Excel faylini yuboring.\n\n"
            "A ustun: Filial kodi\nB ustun: Filial nomi"
        )
        return
    if value == "📊 DASHBOARD":
        await dashboard(message)
        return
    if value == "🏢 FILIALLAR":
        await branches(message)
        return
    if value == "👥 XODIMLAR":
        employees = bot_db.get_all_employees()
        if not employees:
            await message.reply_text("Hozircha xodimlar ro‘yxatdan o‘tmagan.")
            return
        lines = ["👥 XODIMLAR", ""]
        for idx, employee in enumerate(employees[:100], 1):
            lines.append(f"{idx}. {employee['full_name']} — {employee['branch_name']} — {employee['shift']}-smena")
        await message.reply_text("\n".join(lines))
        return
    if value == "🔴 KECHIKKANLAR":
        await status_list(message, "late")
        return
    if value == "❌ KELMAGANLAR":
        await status_list(message, "absent")
        return
    if value == "⚙️ SOZLAMALAR":
        await message.reply_text(
            "⚙️ SOZLAMALAR\n\n"
            "1-smena: kelish 08:15 gacha, ketish 17:00 dan oldin — erta.\n"
            "2-smena: kelish 17:15 gacha, ketish 23:45 dan oldin — erta.\n"
            "2-smena 00:00 gacha bo‘lgan vaqtni keyingi kun sifatida hisoblaydi."
        )
        return
    if value == "📝 So‘rovlar":
        requests = bot_db.list_pending_requests()
        if not requests:
            await message.reply_text("📝 Kutilayotgan so‘rovlar yo‘q.")
            return
        await message.reply_text("\n".join(
            f"{idx}. {row['full_name']} — {row['request_type']}: {row['requested_value']}"
            for idx, row in enumerate(requests, 1)
        ))
        return
    await message.reply_text("Admin menyusidan kerakli tugmani tanlang.", reply_markup=menu_keyboard(ADMIN_MENU))


async def dashboard(message) -> None:
    rows = bot_db.get_today_rows(now_local().date())
    total = len(rows)
    arrived = sum(1 for row in rows if row.get("arrival_at"))
    late = sum(1 for row in rows if row.get("arrival_status") == "LATE")
    absent = total - arrived
    worked = sum(row.get("worked_minutes") or 0 for row in rows)
    await message.reply_text(
        f"📊 DASHBOARD\n\n"
        f"📅 Bugun: {now_local().strftime('%d.%m.%Y')}\n"
        f"👥 Jami xodimlar: {total}\n"
        f"🟢 Kelgan: {arrived}\n"
        f"🔴 Kechikkan: {late}\n"
        f"❌ Kelmagan: {absent}\n"
        f"⏱ Jami ishlangan vaqt: {fmt_duration(worked)}"
    )


async def branches(message) -> None:
    branches_data = bot_db.get_branches()
    if not branches_data:
        await message.reply_text("Hozircha filiallar yuklanmagan.\n“📥 Excel yuklash” tugmasidan foydalaning.")
        return
    buttons = [
        [InlineKeyboardButton(f"🏢 {branch['branch_name']}", callback_data=f"branch_view:{branch['id']}")]
        for branch in branches_data[:80]
    ]
    await message.reply_text("🏢 FILIALLAR:", reply_markup=InlineKeyboardMarkup(buttons))


async def show_branch(message, branch_id: int) -> None:
    branch = bot_db.get_branch(branch_id)
    if not branch:
        await message.reply_text("Filial topilmadi.")
        return
    rows = [row for row in bot_db.get_today_rows(now_local().date()) if row.get("branch_id") == branch_id]
    arrived = sum(1 for row in rows if row.get("arrival_at"))
    late = sum(1 for row in rows if row.get("arrival_status") == "LATE")
    lines = [
        f"🏢 {branch['branch_name']}",
        f"👥 Xodimlar: {len(rows)}",
        f"🟢 Kelgan: {arrived}",
        f"🔴 Kechikkan: {late}",
        "",
    ]
    for row in rows[:50]:
        lines.append(f"{row['full_name']} — {status_text(row.get('arrival_status') or 'ABSENT')} {fmt_time(row.get('arrival_at'))}")
    await message.reply_text("\n".join(lines))


async def status_list(message, kind: str) -> None:
    rows = bot_db.get_today_rows(now_local().date())
    if kind == "late":
        selected = [row for row in rows if row.get("arrival_status") == "LATE"]
        title = "🔴 BUGUN KECHIKKANLAR"
    else:
        selected = [row for row in rows if not row.get("arrival_at")]
        title = "❌ BUGUN KELMAGANLAR"
    if not selected:
        await message.reply_text(f"{title}\n\nRo‘yxat bo‘sh.")
        return
    lines = [title, ""]
    for idx, row in enumerate(selected, 1):
        extra = f"+{row['late_minutes']} daqiqa" if kind == "late" else "Kelmagan"
        lines.append(f"{idx}. 👤 {row['full_name']}\n🏢 {row['branch_name']}\n👔 {row['position']}\n🕐 {fmt_time(row.get('arrival_at'))} — {extra}\n")
    await message.reply_text("\n".join(lines))


async def send_report(message, kind: str) -> None:
    current = now_local().date()
    if kind == "today":
        start = end = current
        label = current.strftime("%d-%m-%Y")
    elif kind == "weekly":
        start = current - timedelta(days=current.weekday())
        end = min(start + timedelta(days=6), current)
        label = f"{start.strftime('%d-%m')} — {end.strftime('%d-%m-%Y')}"
    else:
        start = current.replace(day=1)
        next_month = (start.replace(day=28) + timedelta(days=4)).replace(day=1)
        end = min(next_month - timedelta(days=1), current)
        label = current.strftime("%B_%Y")
    await message.reply_text("⏳ Hisobot tayyorlanmoqda...")
    output = await asyncio.to_thread(build_report, start, end, label)
    filename = f"FILIAL_ATTENDANCE_{label}.xlsx"
    await message.reply_document(document=InputFile(output, filename=filename), caption=f"📎 {filename}")


async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    logger.exception("Telegram update error", exc_info=context.error)


async def post_init(application: Application) -> None:
    await application.bot.set_my_commands([
        BotCommand("start", "Botni boshlash"),
        BotCommand("whoami", "Telegram ID raqamini ko‘rish"),
    ])


def build_application() -> Application:
    token = os.environ.get("TELEGRAM_BOT_TOKEN")
    if not token:
        raise RuntimeError("TELEGRAM_BOT_TOKEN maxfiy sozlamasi topilmadi.")
    bot_db.init_db()
    application = Application.builder().token(token).post_init(post_init).build()
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("whoami", whoami))
    application.add_handler(CallbackQueryHandler(callback))
    application.add_handler(MessageHandler(filters.CONTACT, contact))
    application.add_handler(MessageHandler(filters.Document.ALL, document))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, text))
    application.add_error_handler(error_handler)
    return application


def main() -> None:
    application = build_application()
    logger.info("FILIAL ATTENDANCE bot ishga tushmoqda.")
    application.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()